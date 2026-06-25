import openpyxl, re, json
from openpyxl.utils import get_column_letter

def load(path):
    wv=openpyxl.load_workbook(path, data_only=True)
    wf=openpyxl.load_workbook(path, data_only=False)
    return wv,wf

def num(v): return v if isinstance(v,(int,float)) else 0
def s(v): return str(v).strip() if v is not None else ""

def cellv(ws,r,c): return ws.cell(row=r+1,column=c+1).value  # 0-indexed

def parse(path, p4_label):
    wv,wf=load(path)
    rk=wv['Rekapitulace stavby']
    # rekap period header row idx91, find Prostavěnost headers
    periods=[]
    for c in range(49,150):
        v=s(cellv(rk,91,c))
        if 'Prostav' in v:
            periods.append((v.replace('Prostavěnost ','').strip(), c+1))
    rk_labels=[p[0] for p in periods]; rk_valcols=[p[1] for p in periods]
    objmap={'00':'VRN','03':'03-Int','04':'04-Ext'}
    objrows={'00':94,'03':97,'04':98}
    rk_obj={}  # obj -> (budget, [P1,P2,P3])
    for code,ri in objrows.items():
        b=num(cellv(rk,ri,32))
        ps=[num(cellv(rk,ri,vc)) for vc in rk_valcols]
        rk_obj[objmap[code]]=(b,ps)
    # sheets
    sheets={}
    for n in wf.sheetnames:
        for pfx,lab in objmap.items():
            if n.startswith(pfx): sheets[lab]=n
    # detect period columns + sum rows per sheet
    byObj={}; colmap={}; sumrows={}
    NPdet=len(rk_labels)
    for lab,sn in sheets.items():
        wfs=wf[sn]; wvs=wv[sn]
        sumcells=[]  # (col,row,val)
        for r in range(0,wfs.max_row):
            for c in range(18,46):
                f=cellv(wfs,r,c)
                if isinstance(f,str) and f.startswith('=SUM(') and c!=19:
                    val=num(cellv(wvs,r,c))
                    if abs(val)>1: sumcells.append((c,r,val))
        # map to periods
        budget,rkp=rk_obj[lab]
        cmap={}; srows=set()
        for (c,r,val) in sumcells:
            srows.add(r)
            assigned=False
            for pi in range(len(rkp)):
                if pi not in cmap and abs(val-rkp[pi])<2:
                    cmap[pi]=c; assigned=True; break
            if not assigned:
                cmap[3]=c  # P4 = leftover
        sumrows[lab]=srows; colmap[lab]=cmap
        totals=[]
        for pi in range(4):
            if pi in cmap:
                # total = the SUM value at that col
                tv=[v for (cc,rr,v) in sumcells if cc==cmap[pi]]
                totals.append(round(tv[0]) if tv else 0)
            else: totals.append(0)
        byObj[lab]={'budget':round(budget),'p':totals}
    NP=4
    labels=rk_labels[:3]+[p4_label]
    totP=[sum(byObj[o]['p'][pi] for o in byObj) for pi in range(4)]
    auth={'budget':round(sum(byObj[o]['budget'] for o in byObj)),
          'periods':labels,'totP':totP,'cum':sum(totP),'byObj':byObj}
    # extract items
    items=[]
    for lab,sn in sheets.items():
        wvs=wv[sn]; wfs=wf[sn]
        cmap=colmap[lab]; srows=sumrows[lab]
        colFor=[cmap.get(pi) for pi in range(4)]
        maxr=wvs.max_row
        ck=None
        def push(it):
            if it and (any(x!=0 for x in it['p']) or it['b']>0): items.append(it)
        for r in range(9,maxr):
            if r in srows: continue  # skip SUM-total rows (the bug fix)
            typ=s(cellv(wvs,r,3)); desc=s(cellv(wvs,r,5))[:80]; bv=num(cellv(wvs,r,9))
            pv=[round(num(cellv(wvs,r,c))*100)/100 if c is not None else 0 for c in colFor]
            if typ=='K':
                push(ck); ck={'s':lab,'d':desc,'b':round(bv*100)/100,'p':pv}
            elif typ in ('M','VV') and ck:
                ck['b']=round((ck['b']+bv)*100)/100
                ck['p']=[round((a+b)*100)/100 for a,b in zip(ck['p'],pv)]
            elif any(x!=0 for x in pv):
                push(ck); ck=None
                note=''
                for nc in [24,28,34,38,40,5,4]:
                    if not note:
                        nv=s(cellv(wvs,r,nc))
                        if nv and not nv.startswith('='): note=nv[:80]
                # try formula expression as label hint
                if not note:
                    for c in colFor:
                        if c is not None:
                            fexpr=cellv(wfs,r,c)
                            if isinstance(fexpr,str) and fexpr.startswith('='):
                                note=fexpr[1:][:40]; break
                items.append({'s':lab,'d':note or ('Řádek '+str(r+1)),'b':0,'p':pv})
        push(ck)
    # status
    for it in items:
        c2=sum(it['p'])
        if it['b']==0:
            it['st']='manual' if c2>0 else ('meneprace' if c2<0 else 'zero')
        else:
            pct=c2/it['b']*100
            if any(v<0 for v in it['p'][1:]): it['st']='meneprace'
            elif pct>100: it['st']='viceprace'
            elif pct>=99: it['st']='done'
            elif pct>0: it['st']='partial'
            else: it['st']='zero'
    return auth, items

if __name__=='__main__':
    auth,items=parse('PAS11_8.6.xlsx','k 8.6.')
    print("=== AUTH ===")
    print(" budget:",auth['budget']," cum:",auth['cum'])
    print(" periods:",auth['periods'])
    print(" totP:",auth['totP'])
    for o,v in auth['byObj'].items(): print(f"   {o:7} budget={v['budget']:>9} p={v['p']}")
    print(f"\n=== ITEMS: {len(items)} total ===")
    # reconcile: sum item p per object vs byObj
    print("\n=== RECONCILIATION (item sums vs AUTH) ===")
    for pi in range(4):
        ds=round(sum(it['p'][pi] for it in items))
        print(f"  P{pi+1}: items={ds:>10}  AUTH.totP={auth['totP'][pi]:>10}  diff={ds-auth['totP'][pi]:>+6}")
    for o in auth['byObj']:
        for pi in range(4):
            ds=round(sum(it['p'][pi] for it in items if it['s']==o))
            t=auth['byObj'][o]['p'][pi]
            if abs(ds-t)>1: print(f"    MISMATCH {o} P{pi+1}: items={ds} auth={t} diff={ds-t}")
    # count by status & object
    from collections import Counter
    print("\n status:",dict(Counter(it['st'] for it in items)))
    print(" by obj:",dict(Counter(it['s'] for it in items)))
    json.dump({'auth':auth,'items':items}, open('parsed_8.6.json','w'), ensure_ascii=False)
    print("\n wrote parsed_8.6.json")
